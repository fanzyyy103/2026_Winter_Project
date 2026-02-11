import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# different target varibles
CONDITIONS = [
    "perihilar_infiltrate",
    "pneumonia",
    "bronchitis",
    "interstitial",
    "diseased_lungs",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pulmonary_nodules",
    "pleural_effusion",
    "rtm",
    "focal_caudodorsal_lung",
    "focal_perihilar",
    "pulmonary_hypoinflation",
    "right_sided_cardiomegaly",
    "pericardial_effusion",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "left_sided_cardiomegaly",
    "thoracic_lymphadenopathy",
    "esophagitis",
]

# function to convert 1 to abnormal and 0 to normal
def to_binary(x):
    return 1 if str(x).strip().lower() == "abnormal" else 0

# output the confusion matrix
def write_confusion_matrix(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "condition",
        "True Positive",
        "False Negative",
        "True Negative",
        "False Positive",
        "Sensitivity",
        "Specificity",
        "Check",
        "Positive Ground Truth",
        "Negative Ground Truth",
    ]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    row = 2
    # create the labling by manual and AI
    for cond in CONDITIONS:
        gt_col = f"{cond}_gt"
        llm_col = f"{cond}_llm"

        if gt_col not in df.columns or llm_col not in df.columns:
            raise ValueError(f"Missing columns: {gt_col}, {llm_col}")

        y_true = df[gt_col].apply(to_binary)
        y_pred = df[llm_col].apply(to_binary)

        # calculate the TP,FN,TN,FP
        tp = ((y_true == 1) & (y_pred == 1)).sum()
        fn = ((y_true == 1) & (y_pred == 0)).sum()
        tn = ((y_true == 0) & (y_pred == 0)).sum()
        fp = ((y_true == 0) & (y_pred == 1)).sum()

        # write into confusion matrix and process excel calculation
        ws.cell(row=row, column=1, value=cond)
        ws.cell(row=row, column=2, value=tp)
        ws.cell(row=row, column=3, value=fn)
        ws.cell(row=row, column=4, value=tn)
        ws.cell(row=row, column=5, value=fp)
        
        ws.cell(row=row, column=6, value=f"=IFERROR(B{row}/(B{row}+C{row}),0)")
        ws.cell(row=row, column=7, value=f"=IFERROR(D{row}/(D{row}+E{row}),0)")
        ws.cell(row=row, column=8, value=f"=SUM(B{row}:E{row})")
        ws.cell(row=row, column=9, value=f"=SUM(B{row}:C{row})")
        ws.cell(row=row, column=10, value=f"=SUM(D{row}:E{row})")

        row += 1

    wb.save(output_path)

